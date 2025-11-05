#!/usr/bin/env python3
"""
Tool: Import Text Conversations

Imports text conversation data from Excel into MongoDB participants collection.
Performs residence/demographic matching and creates participant engagement records.

This is a reusable tool for one-time imports/updates of text campaign data.

Usage:
    # Dry run (validation only)
    python scripts/import_text_conversations_tool.py --dry-run --verbose

    # Import conversations
    python scripts/import_text_conversations_tool.py --campaign-id 690b0058eadaad6f0e0dbfb3

    # Import with specific Excel file
    python scripts/import_text_conversations_tool.py --excel data/campaign_texting/custom.xlsx --sheet Conversations
"""
import os
import sys
import json
import re
import csv
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import openpyxl
from pymongo import MongoClient
from dotenv import load_dotenv
from bson import ObjectId

from src.models.participant import Participant, TextEngagement
from src.models.common import ResidenceReference, DemographicReference
from src.models.campaign import Campaign, TextStatistics, CampaignStatCount
from src.tools.residence_matcher import ResidenceMatcher, PhoneNormalizer

load_dotenv()

# Default paths
DEFAULT_EXCEL = Path(__file__).parent.parent / 'data' / 'campaign_texting' / 'Empower_Saves_Texts_All.xlsx'
ZIPCODE_COUNTY_MAP = Path(__file__).parent.parent / 'data' / 'zipcode_to_county_cache.json'


class ConversationImporter:
    """Import text conversations with residence matching"""

    def __init__(self, excel_path: Path, campaign_id: str, dry_run: bool = False, verbose: bool = False, limit: Optional[int] = None):
        self.excel_path = excel_path
        self.campaign_id = campaign_id
        self.dry_run = dry_run
        self.verbose = verbose
        self.limit = limit

        # MongoDB connection
        self.mongo_host = os.getenv('MONGODB_HOST_RM', 'localhost')
        self.mongo_port = int(os.getenv('MONGODB_PORT_RM', '27017'))
        self.mongo_db = os.getenv('MONGODB_DATABASE_RM', 'empower_development')

        print(f"Connecting to MongoDB: {self.mongo_host}:{self.mongo_port}/{self.mongo_db}")
        self.client = MongoClient(self.mongo_host, self.mongo_port)
        self.db = self.client[self.mongo_db]

        # Load zipcode mapping
        with open(ZIPCODE_COUNTY_MAP, 'r') as f:
            data = json.load(f)
            self.zipcode_map = data.get('zipcode_map', {})

        # Statistics
        self.stats = {
            'total_conversations': 0,
            'total_contacts': 0,
            'matched_residence': 0,
            'matched_demographic': 0,
            'no_match': 0,
            'created_participants': 0,
            'updated_participants': 0,
            'match_methods': defaultdict(int),
        }

        # Track unmatched contacts for CSV export
        self.unmatched_contacts = []

    def load_conversations(self, sheet_name: str = 'Conversations') -> Dict[str, List[Dict]]:
        """
        Load conversations from Excel and group by phone number

        Returns:
            Dict mapping phone number to list of conversation records
        """
        print(f"\nLoading conversations from: {self.excel_path}")
        print(f"Sheet: {sheet_name}")

        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Group conversations by phone
        conversations_by_phone = defaultdict(list)

        # Determine max rows to process
        max_row = ws.max_row
        if self.limit:
            max_row = min(max_row, self.limit + 1)  # +1 because row 1 is header
            print(f"LIMIT: Processing first {self.limit} conversation rows only")

        for row_idx in range(2, max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[header] = cell_value

            phone = row_data.get('Phone Number')
            if phone:
                norm_phone = PhoneNormalizer.normalize(phone)
                conversations_by_phone[norm_phone].append(row_data)

            self.stats['total_conversations'] += 1

            if row_idx % 10000 == 0:
                print(f"  Loaded {row_idx} rows...")

        self.stats['total_contacts'] = len(conversations_by_phone)
        print(f"Loaded {self.stats['total_conversations']} conversation records")
        print(f"Unique contacts: {self.stats['total_contacts']}")

        return conversations_by_phone

    def get_county_from_zipcode(self, zipcode) -> Optional[str]:
        """Map zipcode to county"""
        if not zipcode:
            return None
        clean_zip = re.sub(r'\D', '', str(zipcode))[:5]
        return self.zipcode_map.get(clean_zip)

    def match_to_residence(self, phone: str, conversation_data: List[Dict]) -> Tuple[Optional[ResidenceReference], Optional[DemographicReference], str]:
        """
        Match phone number to residence/demographic data using enhanced 8-strategy matching

        Args:
            phone: Normalized phone number
            conversation_data: List of conversation records for this phone

        Returns:
            (residence_ref, demographic_ref, match_method)
        """
        # Get contact info from first conversation record (all should be same)
        first_msg = conversation_data[0]

        street = first_msg.get('Street')
        city = first_msg.get('City')
        state = first_msg.get('State')
        zipcode = first_msg.get('Zipcode')
        county_raw = first_msg.get('County')

        # Determine county
        county = None
        if county_raw:
            county = f"{county_raw}County" if not county_raw.endswith('County') else county_raw
        elif zipcode:
            county = self.get_county_from_zipcode(zipcode)

        if not county:
            if self.verbose:
                print(f"  ⚠️  No county found for phone {phone}")
            return None, None, "no_county"

        # Use enhanced matcher with 8 strategies
        matcher = ResidenceMatcher(self.db, county=county)
        residence_ref, demographic_ref, match_method = matcher.match(
            phone=phone,
            address=street,
            zipcode=zipcode
        )

        # Update statistics
        if match_method == "no_match" or match_method == "collection_not_found":
            self.stats['no_match'] += 1
            if self.verbose:
                print(f"  ❌ No match for phone {phone} in {county}")
        else:
            if demographic_ref:
                self.stats['matched_demographic'] += 1
            if residence_ref:
                self.stats['matched_residence'] += 1
            self.stats['match_methods'][match_method] = self.stats['match_methods'].get(match_method, 0) + 1

            if self.verbose:
                parcel_id = residence_ref.parcel_id if residence_ref else demographic_ref.parcel_id if demographic_ref else "N/A"
                print(f"  ✅ Matched {phone} via {match_method} to {county}/{parcel_id}")

        return residence_ref, demographic_ref, match_method

    def import_conversations(self, conversations_by_phone: Dict[str, List[Dict]]):
        """
        Import conversations into participants collection

        Args:
            conversations_by_phone: Dict mapping phone to conversation records
        """
        print(f"\n{'='*80}")
        print(f"IMPORTING CONVERSATIONS")
        print(f"{'='*80}")
        print(f"Mode: {'DRY RUN' if self.dry_run else 'LIVE IMPORT'}")
        print(f"Campaign ID: {self.campaign_id}")

        participants_coll = self.db['participants']
        campaigns_coll = self.db['campaigns']

        # Load campaign
        campaign_oid = ObjectId(self.campaign_id)
        campaign_doc = campaigns_coll.find_one({'_id': campaign_oid})
        if not campaign_doc:
            raise ValueError(f"Campaign not found: {self.campaign_id}")

        campaign_name = campaign_doc.get('name', 'Unknown')
        print(f"Campaign: {campaign_name}")
        print(f"{'='*80}\n")

        # Statistics for campaign update
        total_sent = 0
        total_delivered = 0
        total_clicked = 0
        total_failed = 0
        total_opt_outs = 0

        for idx, (phone, conversations) in enumerate(conversations_by_phone.items(), 1):
            if idx % 100 == 0:
                print(f"Processing contact {idx}/{self.stats['total_contacts']}...")

            # Match to residence/demographic
            residence_ref, demographic_ref, match_method = self.match_to_residence(phone, conversations)

            # Track unmatched for CSV export
            if match_method in ("no_match", "no_county", "collection_not_found"):
                first_msg = conversations[0]
                self.unmatched_contacts.append({
                    'phone': phone,
                    'street': first_msg.get('Street'),
                    'city': first_msg.get('City'),
                    'state': first_msg.get('State'),
                    'zipcode': first_msg.get('Zipcode'),
                    'county': first_msg.get('County'),
                    'match_method': match_method,
                    'message_count': len(conversations)
                })

            # Create participant
            participant = Participant.from_text_conversation(
                phone=phone,
                campaign_id=self.campaign_id,
                conversation_data=conversations,
                residence_ref=residence_ref,
                demographic_ref=demographic_ref
            )

            # Extract engagement stats for campaign aggregation
            if participant.engagements:
                engagement = participant.engagements[0]  # Should have exactly one
                if isinstance(engagement, TextEngagement):
                    total_sent += engagement.messages_sent
                    total_delivered += engagement.messages_delivered
                    total_failed += engagement.messages_failed
                    # Approximation: clicked if replied
                    if engagement.replied:
                        total_clicked += 1
                    if engagement.opted_out:
                        total_opt_outs += 1

            if not self.dry_run:
                # Check if participant already exists
                existing = participants_coll.find_one({'contact_id': phone})

                if existing:
                    # Update existing participant (add/update engagement)
                    participants_coll.update_one(
                        {'_id': existing['_id']},
                        {'$set': participant.to_mongo_dict()}
                    )
                    self.stats['updated_participants'] += 1
                else:
                    # Insert new participant
                    participants_coll.insert_one(participant.to_mongo_dict())
                    self.stats['created_participants'] += 1
            else:
                self.stats['created_participants'] += 1

        # Update campaign statistics
        if not self.dry_run:
            campaign_stats = TextStatistics(
                sent=CampaignStatCount(unique=len(conversations_by_phone), total=total_sent),
                delivered=CampaignStatCount(unique=total_delivered, total=total_delivered),
                clicked=CampaignStatCount(unique=total_clicked, total=total_clicked),
                failed=CampaignStatCount(unique=total_failed, total=total_failed),
                opt_outs=CampaignStatCount(unique=total_opt_outs, total=total_opt_outs)
            )

            campaigns_coll.update_one(
                {'_id': campaign_oid},
                {'$set': {
                    'statistics': campaign_stats.dict() if hasattr(campaign_stats, 'dict') else campaign_stats.model_dump(),
                    'synced_at': datetime.now()
                }}
            )
            print(f"\n✅ Updated campaign statistics")

    def print_statistics(self):
        """Print import statistics"""
        print(f"\n{'='*80}")
        print(f"IMPORT STATISTICS")
        print(f"{'='*80}")

        print(f"\nConversations:")
        print(f"  Total conversation records: {self.stats['total_conversations']:,}")
        print(f"  Unique contacts:            {self.stats['total_contacts']:,}")

        print(f"\nMatching Results:")
        print(f"  Matched to residence:       {self.stats['matched_residence']:,}")
        print(f"  Matched to demographic:     {self.stats['matched_demographic']:,}")
        print(f"  No match:                   {self.stats['no_match']:,}")

        print(f"\nMatch Methods:")
        for method, count in sorted(self.stats['match_methods'].items()):
            print(f"  {method:20s}: {count:,}")

        print(f"\nParticipants:")
        print(f"  Created:                    {self.stats['created_participants']:,}")
        print(f"  Updated:                    {self.stats['updated_participants']:,}")

        print(f"{'='*80}")

    def write_unmatched_csv(self):
        """Write unmatched contacts to CSV for later processing"""
        if not self.unmatched_contacts:
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_path = Path(__file__).parent.parent / 'data' / 'tmp' / f'unmatched_contacts_{timestamp}.csv'

        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['phone', 'street', 'city', 'state', 'zipcode', 'county', 'match_method', 'message_count']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.unmatched_contacts)

        print(f"\n📄 Wrote {len(self.unmatched_contacts)} unmatched contacts to: {csv_path}")

    def close(self):
        """Close MongoDB connection"""
        self.client.close()


def main():
    """Main execution"""
    import argparse

    parser = argparse.ArgumentParser(description='Import text conversations into MongoDB')
    parser.add_argument('--excel', type=Path, default=DEFAULT_EXCEL, help='Excel file path')
    parser.add_argument('--sheet', default='Conversations', help='Worksheet name')
    parser.add_argument('--campaign-id', required=True, help='MongoDB Campaign _id')
    parser.add_argument('--dry-run', action='store_true', help='Validation only (no writes)')
    parser.add_argument('--verbose', action='store_true', help='Verbose output')
    parser.add_argument('--limit', type=int, help='Limit number of conversation rows to process (for testing)')

    args = parser.parse_args()

    print(f"{'='*80}")
    print(f"TEXT CONVERSATIONS IMPORT TOOL")
    print(f"{'='*80}")
    print(f"Started at: {datetime.now()}")
    print(f"Excel file: {args.excel}")
    print(f"Worksheet: {args.sheet}")
    print(f"Campaign ID: {args.campaign_id}")
    print(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE IMPORT'}")

    importer = ConversationImporter(
        excel_path=args.excel,
        campaign_id=args.campaign_id,
        dry_run=args.dry_run,
        verbose=args.verbose,
        limit=args.limit
    )

    try:
        conversations = importer.load_conversations(args.sheet)
        importer.import_conversations(conversations)
        importer.print_statistics()
        importer.write_unmatched_csv()

        print(f"\nCompleted at: {datetime.now()}")
        return 0

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        importer.close()


if __name__ == '__main__':
    sys.exit(main())
